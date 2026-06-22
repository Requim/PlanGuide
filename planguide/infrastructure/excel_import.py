"""Excel 导入适配器。"""

from __future__ import annotations

from io import BytesIO
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook

from planguide.config import settings
from planguide.domain.plan import PlanItem, PlanTemplate, default_fields


class ExcelImportAdapter:
    def preview(self, filename: str, content: bytes) -> dict[str, Any]:
        self._validate(filename, content)
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        return {"filename": filename, "sheets": self._preview_sheets(workbook)}

    def build_template(self, preview: dict, payload: dict) -> PlanTemplate:
        sheet = self._find_sheet(preview, payload.get("sheet_name", ""))
        mapping = payload.get("mapping", {})
        if not mapping.get("title"):
            raise ValueError("请映射任务标题列")
        items = self._build_items(sheet, mapping)
        if not items:
            raise ValueError("未从 Excel 中解析到有效任务")
        return self._template(preview, payload, items)

    def _validate(self, filename: str, content: bytes):
        if not filename.lower().endswith(".xlsx"):
            raise ValueError("仅支持 .xlsx 文件")
        if len(content) > settings.upload_max_bytes:
            raise ValueError("文件超过上传大小限制")

    def _preview_sheets(self, workbook) -> list[dict[str, Any]]:
        result = []
        for ws in workbook.worksheets:
            rows = list(ws.iter_rows(values_only=True, max_row=settings.excel_max_rows + 1))
            result.append(_sheet_preview(ws.title, rows, ws.max_row))
        return result

    def _find_sheet(self, preview: dict, sheet_name: str) -> dict:
        for sheet in preview.get("sheets", []):
            if sheet.get("name") == sheet_name:
                return sheet
        raise ValueError("未找到指定 sheet")

    def _build_items(self, sheet: dict, mapping: dict) -> list[PlanItem]:
        headers = sheet.get("headers", [])
        rows = sheet.get("rows", [])[: settings.excel_max_rows]
        return [item for item in (_row_to_item(i, headers, row, mapping) for i, row in enumerate(rows, 1)) if item]

    def _template(self, preview: dict, payload: dict, items: list[PlanItem]) -> PlanTemplate:
        title = payload.get("title") or preview.get("filename", "Excel 导入计划")
        return PlanTemplate(
            template_id=f"excel-{uuid4().hex[:12]}",
            title=title,
            description="由 Excel 引导式映射导入",
            source_type="excel",
            items=items,
            fields=default_fields(),
        )


def _sheet_preview(name: str, rows: list[tuple], max_row: int) -> dict[str, Any]:
    headers = [str(value or "").strip() for value in (rows[0] if rows else [])]
    data_rows = [_normalize_row(row) for row in rows[1:]]
    return {
        "name": name,
        "headers": headers,
        "rows": data_rows,
        "sample_rows": data_rows[:7],
        "total_rows": max_row,
    }


def _row_to_item(index: int, headers: list[str], row: list[Any], mapping: dict) -> PlanItem | None:
    title = _cell(row, headers, mapping.get("title"))
    if not str(title or "").strip():
        return None
    return PlanItem(
        item_id=f"excel-row-{index}",
        title=str(title).strip(),
        detail=str(_cell(row, headers, mapping.get("detail")) or ""),
        date=str(_cell(row, headers, mapping.get("date")) or ""),
        group=str(_cell(row, headers, mapping.get("group")) or ""),
        module=str(_cell(row, headers, mapping.get("module")) or ""),
        output=str(_cell(row, headers, mapping.get("output")) or ""),
    )


def _normalize_row(row: tuple) -> list[Any]:
    return [_format_cell(value) for value in row]


def _cell(row: list[Any], headers: list[str], name: str | None) -> Any:
    if not name or name not in headers:
        return None
    index = headers.index(name)
    return row[index] if index < len(row) else None


def _format_cell(value: Any) -> Any:
    if hasattr(value, "date"):
        return value.date().isoformat()
    return value
